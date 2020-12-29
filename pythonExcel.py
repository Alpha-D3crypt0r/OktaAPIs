import openpyxl

wb=openpyxl.load_workbook("GetUserSheet.xlsx")
sheet=wb.active
cell=sheet.cell(1,1)
cellput=sheet.cell(1,2)
cellput.value="Writing Bitch"
print(cell.value)
wb.save("GetUserSheet.xlsx")
