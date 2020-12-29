import requests,json
from termcolor import colored
from colorama import Fore
from colorama import Style
import openpyxl
loc = ("OktaGetUsers.xlsx")
wb=openpyxl.load_workbook(loc)
sheet=wb.active
rows=sheet.max_row
column=sheet.max_column
cell=sheet.cell(1,column+1)
cell.value="Error/Remarks"
wb.save(loc)

print(f"{Fore.GREEN}Kraken's API Get Request Okta Script{Style.RESET_ALL}")
print("Total Records to be processed: ",rows-1)
#Reqests library handles the API calls
tenent = "dev-587547.okta"
token = "002x_RGs87sZcSVrsSWI9haeokHyCH1TjqUL1vgQ1L"
CurrentCellLocation=0
def main():

    for CurrentCellLocation in range(rows-1):
        cellget=sheet.cell(CurrentCellLocation+2,1)
        content=getUser(cellget.value,CurrentCellLocation)
        if content == "Not Found":
            cellput=sheet.cell(CurrentCellLocation+2,column+1)
            cellput.value="User is not found in the database"
            wb.save(loc)
        elif content == "Invalid Token":
            cellput=sheet.cell(CurrentCellLocation+2,column+1)
            cellput.value="API Token is either invalid or is expired. Please create a new token in Okta under Security->API->Tokens"
            wb.save(loc)
        else:
            cellputid=sheet.cell(CurrentCellLocation+2,2)
            cellputid.value=content['id']
            #print("The UserID for {} is {}: ".format(cellget.value,content['id']))

            cellputfirstName=sheet.cell(CurrentCellLocation+2,3)
            cellputfirstName.value=content['profile']['firstName']
            #print("The UserID for {} is {}: ".format(cellget.value,content['profile']['firstName']))

            cellputLastName=sheet.cell(CurrentCellLocation+2,4)
            cellputLastName.value=content['profile']['lastName']
            #print("The UserID for {} is {}: ".format(cellget.value,content['profile']['lastName']))
            print("{} Records are processed.".format(CurrentCellLocation+1))
            wb.save(loc)
#GET USER FUNCTION
def getUser(login,CurrentCellLocation):
    my_headers = {"Authorization":"SSWS"+token, "Accept":"application/json", "Content-Type":"application/json"}
    response=requests.get('https://{}.com/api/v1/users?filter=profile.login%20eq%20"{}"'.format(tenent,login) , headers=my_headers)
    #print(r.text) r.text contains the real response fromt the API Server. r only contains the response type eg. 200 404
    #print(r)  #Print response
    if response.status_code == requests.codes['ok']:
        result=json.loads(response.text) #Converts the parsed response in the form of JSON
        try:
            return result[0] #Because the response is a list of dictionaries hence taking the first dict from the list
        except IndexError:
            #print(f"{Fore.GREEN}The username is either not correct or the user is not found in the OKTA Database{Style.RESET_ALL}")
            print("{} Records are processed.".format(CurrentCellLocation+1))
            code="Not Found"
            return code
    elif response.status_code == 401:
        #print(f'{Fore.GREEN}Check your API Token its not correct{Style.RESET_ALL}')
        print("{} Records are processed.".format(CurrentCellLocation+1))
        code="Invalid Token"
        return code

main()
